# coding=utf-8
# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
from openpyxl import Workbook
from datetime import datetime

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print("Hi, {0}".format(name))  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    now = datetime.now()
    date_time_print = now.strftime("%m_%d_%Y_%H-%M-%S")
    name_report = 'Report_' + date_time_print + '.csv'
    wb = Workbook()  # workBook
    ws = wb.active  # workSheet
    # ws.title("Report Immobiliare")
    # ws = wb[name_report]
    index = 0
    dataImmobile = {'name':[1,2,3],'price':[11,22,33],'space':[111,222,333]}
    # len(dataImmobile.get('name'))
    headers = ['name', 'price', 'space']
    for row in ws.iter_rows(min_row=1, max_col=len(headers), max_row=1):
        counter = 0
        for cell in row:
            cell.value = headers[counter]
            counter = counter + 1


        for row in range(2, 1 + len(dataImmobile.get('name'))):
            counter = 0
            for col in range(1, len(headers)):
                if (ws.cell(row=1, column=col).value == 'price'):
                    prize = dataImmobile.get('price')[counter]
                    ws.cell(row=row, column=col).value = prize
                if (ws.cell(row=1, column=col).value == 'space'):
                    spc = dataImmobile.get('space')[counter]
                    ws.cell(row=row, column=col).value = spc
                if(ws.cell(row=1, column=col).value == 'name'):
                    noum = dataImmobile.get('name')[counter]
                    ws.cell(row=row, column=col).value = noum
                counter = counter + 1

    wb.save(name_report)

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
