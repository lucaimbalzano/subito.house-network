# coding=utf-8
# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
from openpyxl import Workbook
from datetime import datetime

from openpyxl.pivot.table import Reference


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print("Hi, {0}".format(name))  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    now = datetime.now()
    date_time_print = now.strftime("%m_%d_%Y_%H-%M-%S")
    name_report = 'Report_' + date_time_print + '.csv'
    wb = Workbook()  # workBook
    ws = wb.active  # workSheet
    ws.title = "Report Immobiliare"
    index = 0
    dataImmobile = {'name':[1,2,3],'price':[11,22,33],'space':[111,222,333]}

    headers = ['name', 'price', 'space']
    for row in ws.iter_rows(min_row=1, max_col=len(headers), max_row=1):
        counter = 0
        for cell in row:
            cell.value = headers[counter]
            counter = counter + 1

    counters = [0, 0, 0]
    for row_counter in range(1, len(dataImmobile.get('name'))+1, 1):
        print("row: "+str(row_counter))
        for count in range(0,len(dataImmobile.get('name')),1):
            print("col: "+str(count))
            if (headers.index('name') == count):
                temp_name = dataImmobile.get('name')[counters[0]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_name
                counters[0] += 1
            if(headers.index('price') == count):
                temp_price = dataImmobile.get('price')[counters[1]]
                ws.cell(row=row_counter+1, column=count+1).value = temp_price
                counters[1] += 1
            if (headers.index('space') == count):
                temp_space = dataImmobile.get('space')[counters[2]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[2] += 1

    wb.save(name_report)

