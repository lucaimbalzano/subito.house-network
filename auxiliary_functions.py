import datetime as dt
from bs4 import BeautifulSoup
# import pandas as pd
import openpyxl
import requests,re,time
from data_immobile import Data_immobile
from datetime import datetime
from openpyxl import Workbook



def sistema_data_ora(caricato_il):
    if caricato_il == None:
        return None
    if "Oggi" in caricato_il.text:
        index = (caricato_il.text.find(":"))
        ora = int(caricato_il.text[index - 2:index])
        minuti = int(caricato_il.text[index + 1:index + 3])
        orario = dt.time(ora, minuti)
        today = dt.date.today()
        data_con_orario = dt.datetime.combine(today, orario)
    elif "Ieri" in caricato_il.text:
        index = (caricato_il.text.find(":"))
        yesterday = dt.date.today() - dt.timedelta(days=1)
        ora = int(caricato_il.text[index - 2:index])
        minuti = int(caricato_il.text[index + 1:index + 3])
        orario = dt.time(ora, minuti)
        data_con_orario = dt.datetime.combine(yesterday, orario)
    else:
        mesi = ["gen", "feb", "mar", "apr", "mag", "giu", "lug", "ago", "set", "ott", "nov", "dic"]
        split = caricato_il.text.split(" ")
        strorario = split[3]
        index = (strorario.find(":"))
        ora = int(strorario[index - 2:index])
        minuti = int(strorario[index + 1:index + 3])
        orario = dt.time(ora, minuti)
        giorno = int(split[0])
        mese = mesi.index(split[1]) + 1
        data = dt.date(2020, mese, giorno)
        data_con_orario = dt.datetime.combine(data, orario)
    return data_con_orario


def writeExcelByDf(df):
        now = datetime.now()
        # date_time_print = now.strftime("%m_%d_%Y_%H-%M-%S")
        # name_report = 'Report_'+date_time_print+'.xlsx'
        # path = 'C:/Users/lucai/PycharmProjects/New folder/subitoBot/reports/'
        # df = pd.DataFrame(column=[title,number,name,price,space, rooms, floor, description,url_list])
        # with pd.ExcelWriter(path) as writer:
        #     writer.book = openpyxl.load_workbook(path)
        #     df.to_excel(writer, sheet_name='report')
        #     df.to_excel(name_report, index=False, header=True)


def writeExcelByDataImmobile(dataImmobile):
    now = datetime.now()
    date_time_print = now.strftime("%m_%d_%Y_%H-%M-%S")
    path = './reports/'
    name_report = path + 'Report_' + date_time_print + '.xlsx'
    wb = Workbook() #workBook
    ws = wb.active #workSheet
    ws.title = "Report Immobiliare"
    headers = ['name','price','space','rooms','floor','description','title','url','number']

    for row in ws.iter_rows(min_row=1, max_col=len(headers), max_row=1):
        counter = 0
        for cell in row:
            cell.value = headers[counter]
            counter = counter + 1

    counters = [0,0,0,0,0,0,0,0,0]
    for row_counter in range(1, len(dataImmobile.name)+1, 1):
        for count in range(0,len(headers),1):
            if (headers.index('name') == count):
                temp_name = dataImmobile.name[counters[0]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_name
                counters[0] += 1
            if(headers.index('price') == count):
                temp_price = dataImmobile.price[counters[1]]
                ws.cell(row=row_counter+1, column=count+1).value = temp_price
                counters[1] += 1
            if (headers.index('space') == count):
                temp_space = dataImmobile.space[counters[2]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[2] += 1
            if (headers.index('rooms') == count):
                temp_space = dataImmobile.rooms[counters[3]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[3] += 1
            if (headers.index('floor') == count):
                temp_space = dataImmobile.floor[counters[4]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[4] += 1
            if (headers.index('description') == count):
                temp_space = dataImmobile.description[counters[5]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[5] += 1
            if (headers.index('title') == count):
                temp_space = dataImmobile.title[counters[6]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[6] += 1
            if (headers.index('url') == count):
                temp_space = dataImmobile.url[counters[7]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[7] += 1
            if (headers.index('number') == count):
                temp_space = dataImmobile.number[counters[8]]
                ws.cell(row=row_counter + 1, column=count + 1).value = temp_space
                counters[8] += 1

    wb.save(name_report)




def excel_date(df):
    new_date_up = []
    new_imm = []
    for iel in range(len(df)):
        if df.loc[iel]["rooms"] != None:
            new_imm.append(df.loc[iel]["rooms"].strftime("%x %X"))
        else:
            new_imm.append(None)
        new_date_up.append(df.loc[iel]["Data_upload"].strftime("%x %X"))
    df["Data_upload"] = new_date_up
    df["rooms"] = new_imm
    return df


def starting_time(hour, minute):
    if dt.datetime.now().hour != hour or dt.datetime.now().minute != minute:
        today = dt.datetime.today()
        domani = today + dt.timedelta(days=1)
        ora = dt.time(hour, minute)
        until = domani.combine(domani, ora)
        diff = until - dt.datetime.today()
        secnd = diff.total_seconds()
        print("i will start at:", dt.datetime.today() + dt.timedelta(seconds=secnd))
        time.sleep(secnd)